from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from quote_generator import Quotation


def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(ws, row, col, value, bg="0F0F0F", fg="FFFFFF", bold=True, size=10, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = thin_border("888888")
    return c


def dat(ws, row, col, value, bg="FFFFFF", bold=False, color="000000", align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, color=color, size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


def export_excel(quote: Quotation) -> bytes:
    wb = Workbook()

    # ── Sheet 1: Quotation ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Quotation"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    r = 1
    ws.merge_cells(f"A{r}:F{r}")
    t = ws.cell(row=r, column=1, value=f"QUOTATION  —  {quote.quote_number}")
    t.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="0F0F0F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36

    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(row=r, column=1, value=f"From: {quote.sender_company}").font = Font(name="Calibri", size=10, color="444444")
    ws.merge_cells(f"D{r}:F{r}")
    ws.cell(row=r, column=4, value=f"Date: {quote.date.strftime('%d.%m.%Y')}  |  Valid Until: {quote.valid_until.strftime('%d.%m.%Y')}").font = Font(name="Calibri", size=10, color="444444")
    ws.row_dimensions[r].height = 18

    r += 2
    # Line items header
    for ci, h in enumerate(["Pos.", "Description", "Qty.", "Unit", "Unit Price (€)", "Total (€)"], 1):
        hdr(ws, r, ci, h)
    ws.row_dimensions[r].height = 22

    r += 1
    first_data_row = r
    for item in quote.line_items:
        dat(ws, r, 1, item.pos, align="center", bg="F9F8F6")
        dat(ws, r, 2, item.description, bg="F9F8F6")
        dat(ws, r, 3, item.qty, align="right", bg="F9F8F6")
        dat(ws, r, 4, item.unit, align="center", bg="F9F8F6")
        dat(ws, r, 5, item.unit_price, align="right", bg="F9F8F6", num_fmt='€#,##0.00')
        dat(ws, r, 6, item.total, align="right", bg="F9F8F6", bold=True, num_fmt='€#,##0.00')
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    # Totals
    totals = [
        ("Assembly Net",                     quote.net_assembly,    "F9F8F6", "000000", False),
        (f"Margin ({quote.margin_percent:.0f}%)", quote.margin_amount, "F9F8F6", "000000", False),
        ("Net Price",                         quote.net_with_margin, "EEF3FF", "1A4FD6", True),
        (f"VAT ({quote.vat_percent:.0f}%)",  quote.vat_amount,      "F9F8F6", "000000", False),
        ("QUOTATION TOTAL",                   quote.grand_total,     "1A4FD6", "FFFFFF", True),
    ]
    for label, value, bg, fg, bold in totals:
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Calibri", bold=bold, color=fg, size=10 if not bold else 11)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = thin_border()
        for col in [2, 3, 4]:
            cc = ws.cell(row=r, column=col)
            cc.fill = PatternFill("solid", start_color=bg)
            cc.border = thin_border()
        vc = ws.cell(row=r, column=5, value=value)
        vc.font = Font(name="Calibri", bold=bold, color=fg, size=10 if not bold else 11)
        vc.fill = PatternFill("solid", start_color=bg)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.number_format = '€#,##0.00'
        vc.border = thin_border()
        ws.merge_cells(f"E{r}:F{r}")
        ws.row_dimensions[r].height = 20
        r += 1

    r += 2
    ws.cell(row=r, column=1, value=f"Payment Terms: {quote.payment_terms}").font = Font(name="Calibri", size=9, color="666666", italic=True)

    # ── Sheet 2: BOM Mismatch Check ───────────────────────────────────────────
    ws2 = wb.create_sheet("BOM Mismatch Check")
    for col, w in zip("ABCDEFG", [8, 42, 14, 14, 12, 16, 45]):
        ws2.column_dimensions[col].width = w

    r2 = 1
    ws2.merge_cells(f"A{r2}:G{r2}")
    t2 = ws2.cell(row=r2, column=1, value="BOM REFERENCE vs QTY MISMATCH CHECK")
    t2.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t2.fill = PatternFill("solid", start_color="0F0F0F")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r2].height = 30

    r2 += 1
    for ci, h in enumerate(["Pos.", "Reference(s)", "Declared Qty", "Ref Count", "Diff", "Status", "Description"], 1):
        hdr(ws2, r2, ci, h)
    ws2.row_dimensions[r2].height = 22

    r2 += 1
    for row in quote.bom_rows:
        diff = row.actual_ref_count - row.declared_qty
        ok = row.qty_match
        bg = "F2FFF4" if ok else "FFF2F2"
        dat(ws2, r2, 1, row.pos, bg=bg, align="center")
        dat(ws2, r2, 2, row.reference, bg=bg)
        dat(ws2, r2, 3, row.declared_qty, bg=bg, align="center")
        dat(ws2, r2, 4, row.actual_ref_count, bg=bg, align="center")
        dat(ws2, r2, 5, diff if diff != 0 else "-", bg=bg, align="center",
            bold=not ok, color="9C0006" if not ok else "000000")
        sc = ws2.cell(row=r2, column=6, value="✓ MATCH" if ok else "✗ MISMATCH")
        sc.font = Font(name="Calibri", bold=True, color="276221" if ok else "9C0006", size=10)
        sc.fill = PatternFill("solid", start_color="C6EFCE" if ok else "FFC7CE")
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.border = thin_border()
        dat(ws2, r2, 7, row.description[:60], bg=bg)
        ws2.row_dimensions[r2].height = 18
        r2 += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
